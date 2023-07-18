from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from pathlib import Path
import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active

Path = Path(r"C:\Users\name\Projects\fun\plpsBLR.txt")
column_names = ["request_path", "absolute_path", "title", "description", "keywords", "heading"]
for col_num, col_name in enumerate(column_names, start=1):
    sheet.cell(row=1, column=col_num).value = col_name

browser = webdriver.Chrome()
browser.get(
    "https://our_backend")

browser.maximize_window()

sleep(30)
wait = WebDriverWait(browser, 3)
def write_to_excel(row, values):
    for col_num, value in enumerate(values, start=1):
        sheet.cell(row=row, column=col_num).value = value

with open(Path, "r") as file:
    lines = file.readlines()
    for index, line in enumerate(lines, start=2):
        filter1 = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "action-default")))
        filter1.click()
        sleep(1)

        path = wait.until(EC.presence_of_element_located((By.NAME, "request_path")))
        path.send_keys(line)
        path.send_keys(Keys.ENTER)
        sleep(1)

        choose = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[2]/main/div[3]/div/div/div/div[4]/table/tbody/tr/td[7]/div")))
        choose.click()
        sleep(1)
        modify = wait.until(EC.presence_of_element_located(
            (By.XPATH, "/html/body/div[2]/main/div[3]/div/div/div/div[4]/table/tbody/tr/td[7]/div/ul/li[1]/a")))
        modify.click()
        sleep(2)

        request_path = wait.until(EC.presence_of_element_located(
            (By.XPATH, "/html/body/div[2]/main/div[3]/div/div/div/div[2]/div[1]/div[2]/fieldset/div[4]/div[2]/input")))
        request_attr_value = request_path.get_attribute('value')
        sleep(1)

        absolute_path = wait.until(EC.presence_of_element_located(
            (By.XPATH, "/html/body/div[2]/main/div[3]/div/div/div/div[2]/div[1]/div[2]/fieldset/div[5]/div[2]/input")))
        absolute_attr_value = absolute_path.get_attribute('value')
        sleep(1)

        title = wait.until(EC.presence_of_element_located(
            (By.XPATH,
             "/html/body/div[2]/main/div[3]/div/div/div/div[2]/div[2]/div[2]/fieldset/div[1]/div[2]/textarea")))
        title_attribute_value = title.get_attribute('value')
        sleep(1)

        description = wait.until(EC.presence_of_element_located(
            (By.XPATH,
             "/html/body/div[2]/main/div[3]/div/div/div/div[2]/div[2]/div[2]/fieldset/div[2]/div[2]/textarea")))
        description_attr_value = description.get_attribute('value')
        sleep(1)

        key_words = wait.until(EC.presence_of_element_located((By.XPATH,
            "/html/body/div[2]/main/div[3]/div/div/div/div[2]/div[2]/div[2]/fieldset/div[3]/div[2]/textarea")))
        keyWords_attr_value = key_words.get_attribute('value')
        sleep(1)

        heading = wait.until(EC.presence_of_element_located((By.NAME, "heading")))
        heading_attr_value = heading.get_attribute('value')
        sleep(2)

        browser.back()
        sleep(4)
        values = [request_attr_value, absolute_attr_value, title_attribute_value, description_attr_value,
                  keyWords_attr_value, heading_attr_value]
        write_to_excel(index, values)
        workbook.save(r"C:\Users\name\Projects\fun\infoBLR_Metas.xlsx")
        index += 1

workbook.save(r"C:\Users\name\Projects\fun\infoBLR_Metas.xlsx")
browser.quit()
